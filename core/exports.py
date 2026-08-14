"""Geração de relatórios em Excel (.xlsx).

Fica fora do views.py de propósito: aquele ficheiro já passa das 8 mil linhas,
e montagem de planilha é trabalho de formatação, não de view.

O relatório de alunos substituiu a exportação em CSV que era feita no
navegador. Aquela levava 7 campos dos cerca de 20 que o modelo guarda, sem
formatação nenhuma, e o Excel abria tudo como texto.
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Student

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Mesma paleta da identidade do produto: tinta escura e papel.
_TINTA = "171310"
_PAPEL = "FAF7F2"
_LINHA = "E3DCD1"
_ALERTA = "B45309"

_FORMATO_MOEDA = 'R$ #,##0.00'
_FORMATO_DATA = "DD/MM/YYYY"


def _texto_escolha(instancia, campo, choices):
    """Converte o valor guardado ('active') no rótulo em português ('Ativo')."""
    valor = getattr(instancia, campo, "") or ""
    return dict(choices).get(valor, valor)


def _nome_professor(user):
    if not user:
        return ""
    nome = (user.get_full_name() or "").strip()
    return nome or user.username


def _aulas_restantes(aluno):
    """Quantas aulas faltam no pacote. Vazio quando o plano não tem total."""
    total = aluno.lessons_total or 0
    if not total:
        return None
    return max(total - (aluno.lessons_done or 0), 0)


# (título, função que extrai o valor, formato de número, largura da coluna)
_COLUNAS = [
    ("Nome", lambda a: a.name, None, 28),
    ("Status", lambda a: _texto_escolha(a, "status", Student.STATUS_CHOICES), None, 12),
    ("Nível", lambda a: a.level or "", None, 8),
    ("Responsável", lambda a: a.guardians or "", None, 24),
    ("Telefone", lambda a: a.phone or "", None, 18),
    ("E-mail", lambda a: a.email or "", None, 28),
    ("Endereço", lambda a: a.address or "", None, 32),
    ("Tipo de cobrança", lambda a: _texto_escolha(a, "billing_type", Student.BILLING_TYPE_CHOICES), None, 18),
    ("Plano", lambda a: a.plan_name or "", None, 22),
    ("Valor mensal", lambda a: a.monthly_amount, _FORMATO_MOEDA, 14),
    ("Valor por aula", lambda a: a.per_lesson_amount, _FORMATO_MOEDA, 14),
    ("Início do plano", lambda a: a.plan_start_date, _FORMATO_DATA, 15),
    ("Aulas do plano", lambda a: a.lessons_total or 0, "0", 13),
    ("Aulas realizadas", lambda a: a.lessons_done or 0, "0", 15),
    ("Aulas restantes", _aulas_restantes, "0", 14),
    ("Dia de vencimento", lambda a: a.default_due_day, "0", 16),
    ("Forma de pagamento", lambda a: _texto_escolha(a, "preferred_payment_method", Student.PAYMENT_METHOD_CHOICES), None, 18),
    ("Chave Pix", lambda a: a.pix_key or "", None, 24),
    ("Professor parceiro", lambda a: _nome_professor(a.assigned_teacher), None, 22),
    ("Observações", lambda a: (a.teacher_notes or "").strip(), None, 40),
    ("Cadastrado em", lambda a: timezone.localtime(a.created_at).date() if a.created_at else None, _FORMATO_DATA, 14),
]


def _normaliza(valor):
    """openpyxl não aceita Decimal nem datetime com fuso."""
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, datetime):
        return timezone.localtime(valor).replace(tzinfo=None)
    return valor


def _escreve_cabecalho(ws):
    fundo = PatternFill("solid", fgColor=_TINTA)
    fonte = Font(bold=True, color=_PAPEL, size=11)
    alinhamento = Alignment(vertical="center", horizontal="left")
    for coluna, (titulo, _, _, largura) in enumerate(_COLUNAS, start=1):
        celula = ws.cell(row=1, column=coluna, value=titulo)
        celula.fill = fundo
        celula.font = fonte
        celula.alignment = alinhamento
        ws.column_dimensions[get_column_letter(coluna)].width = largura
    ws.row_dimensions[1].height = 26


def planilha_alunos(alunos, nome_professor=""):
    """Monta o .xlsx do relatório de alunos e devolve os bytes.

    `alunos` deve vir na ordem em que serão escritos: quem chama é responsável
    por respeitar a ordenação que o utilizador vê na tela.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"

    _escreve_cabecalho(ws)

    zebra = PatternFill("solid", fgColor=_PAPEL)
    borda = Border(bottom=Side(style="thin", color=_LINHA))
    alinhamento = Alignment(vertical="top", wrap_text=False)

    for indice, aluno in enumerate(alunos):
        linha = indice + 2
        for coluna, (_, extrair, formato, _largura) in enumerate(_COLUNAS, start=1):
            celula = ws.cell(row=linha, column=coluna, value=_normaliza(extrair(aluno)))
            if formato:
                celula.number_format = formato
            celula.alignment = alinhamento
            celula.border = borda
            if indice % 2:
                celula.fill = zebra

    # Destaca em âmbar quem está com o pacote acabando (2 aulas ou menos),
    # que é a mesma regra do lembrete de fim de pacote.
    coluna_restantes = [t for t, *_ in _COLUNAS].index("Aulas restantes") + 1
    for indice, aluno in enumerate(alunos):
        restantes = _aulas_restantes(aluno)
        if restantes is not None and restantes <= 2 and aluno.status == Student.STATUS_ACTIVE:
            celula = ws.cell(row=indice + 2, column=coluna_restantes)
            celula.font = Font(bold=True, color=_ALERTA)

    # Primeira linha sempre visível ao rolar, e filtro em todas as colunas.
    ws.freeze_panes = "A2"
    if alunos:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUNAS))}{len(alunos) + 1}"

    _escreve_resumo(wb, alunos, nome_professor)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _escreve_resumo(wb, alunos, nome_professor):
    """Segunda aba com os números que o professor olharia primeiro."""
    ws = wb.create_sheet("Resumo")
    negrito = Font(bold=True, size=11)
    titulo = Font(bold=True, size=14)

    ws["A1"] = "Relatório de alunos"
    ws["A1"].font = titulo
    ws["A2"] = f"Professor: {nome_professor}" if nome_professor else ""
    ws["A3"] = f"Gerado em: {timezone.localtime().strftime('%d/%m/%Y às %H:%M')}"

    ativos = [a for a in alunos if a.status == Student.STATUS_ACTIVE]
    mensalidades = sum(float(a.monthly_amount) for a in ativos if a.monthly_amount)
    acabando = [a for a in ativos if (_aulas_restantes(a) or 99) <= 2]

    linhas = [
        ("Total de alunos", len(alunos), None),
        ("Ativos", len(ativos), None),
        ("Pausados", len([a for a in alunos if a.status == Student.STATUS_PAUSED]), None),
        ("Encerrados", len([a for a in alunos if a.status == Student.STATUS_ENDED]), None),
        ("Receita mensal fixa dos ativos", mensalidades, _FORMATO_MOEDA),
        ("Ativos com pacote acabando (2 aulas ou menos)", len(acabando), None),
    ]
    for indice, (rotulo, valor, formato) in enumerate(linhas, start=5):
        ws.cell(row=indice, column=1, value=rotulo).font = negrito
        celula = ws.cell(row=indice, column=2, value=valor)
        if formato:
            celula.number_format = formato

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
